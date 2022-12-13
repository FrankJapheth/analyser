from openpyxl import workbook, load_workbook
from openpyxl.workbook import Workbook


def format_property(property_value: str, formatters: list) -> str:
    formatted_value: str = ''

    if 'rmvwp' in formatters:
        formatted_value = property_value.replace(' ', '')

    if 'rmvftp' in formatters and formatted_value == '':
        formatted_value = property_value.replace('.', '')
    elif 'rmvftp' in formatters and formatted_value != '':
        formatted_value = formatted_value.replace('.', '')

    if 'mksml' in formatters and formatted_value == '':
        formatted_value = property_value.lower()
    elif 'mksml' in formatters and formatted_value != '':
        formatted_value = formatted_value.lower()

    if 'rmvnl' in formatters and formatted_value == '':
        formatted_value = property_value.replace('\n', '')
    elif 'rmvnl' in formatters and formatted_value != '':
        formatted_value = formatted_value.replace('\n', '')

    return formatted_value


def find_values(list_a: list, list_b: list) -> dict:
    present_props = []
    props_not_present = []
    for product_prop in list_a:
        if product_prop not in list_b:
            props_not_present.append({
                'name': product_prop,
                'index_in_a': list_a.index(product_prop),
                'index_in_b': -1
            })
        else:
            present_props.append({
                'name': product_prop,
                'index_in_a': list_a.index(product_prop),
                'index_in_b': list_b.index(product_prop)
            })

    return {
        'properties_present': present_props,
        'properties_not_present': props_not_present
    }


def get_redundancy(products_list: list):
    my_list = sorted(products_list)

    duplicates = []
    for i in my_list:
        if my_list.count(i) > 1:
            if i not in duplicates:
                duplicates.append(i)

    print(len(duplicates))


class DalaliProducts:
    PRODUCTS_FILE: workbook
    PRODUCTS_SHEET = ''

    def read_products_file(self, path_to_file: str) -> workbook:
        self.PRODUCTS_FILE = load_workbook(filename=path_to_file)

        return self.PRODUCTS_FILE

    def get_products_details(self, sheet_name: str, max_col: int, max_row: int, header_row: int = 1) -> dict:
        sheet_details = self.PRODUCTS_FILE[sheet_name]
        property_row = 1
        headers = []
        products_details = []
        for row in sheet_details.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
            product_details = {}
            if property_row == header_row:
                for cell in row:
                    headers.append(cell.value)
            else:
                for cell in row:
                    product_details[headers[row.index(cell)]] = cell.value
                products_details.append(product_details)
            property_row += 1

        sheet_data = {
            'headers': headers,
            'data': products_details
        }

        return sheet_data

    def create_workbook(self):

        self.PRODUCTS_FILE = Workbook()

        self.PRODUCTS_SHEET = self.PRODUCTS_FILE.active

        return self.PRODUCTS_SHEET

    def write_data(self, data, path_to_file: str):

        data_index = 1
        alph_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                        'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                        'U', 'V', 'W', 'X', 'Y', 'Z']
        for sheet_data in data:
            if data_index == 1:
                keys_index = 0
                for data_key in sheet_data.keys():
                    self.PRODUCTS_SHEET[
                        f'{alph_letters[keys_index]}{data_index}'
                    ] = data_key
                    keys_index += 1
            data_index += 1
            keys_index = 0
            for data_key in sheet_data.keys():
                self.PRODUCTS_SHEET[
                    f'{alph_letters[keys_index]}{data_index}'
                ] = sheet_data[data_key]
                keys_index += 1

        self.savefile(path_to_file)

    def savefile(self, file_path: str):
        self.PRODUCTS_FILE.save(file_path)
